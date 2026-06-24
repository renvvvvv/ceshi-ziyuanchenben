#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""偏差报告生成 — 含数量+天数+人天/台天 (2026.6.18)"""
import json, sys, os, math
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import resource_plan as rp

with open(os.path.join(os.path.dirname(__file__), '..', 'config_v100.json'), encoding='utf-8') as f:
    cfg = json.load(f)

import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

thin = Side(style='thin')
border = Border(left=thin,right=thin,top=thin,bottom=thin)
hfont = Font(bold=True, size=11)
tfont = Font(bold=True, size=14)
rfill = PatternFill(start_color='FCE4D6',end_color='FCE4D6',fill_type='solid')
yfill = PatternFill(start_color='FFF2CC',end_color='FFF2CC',fill_type='solid')

def pct(v1,v2):
    try: return 0.0 if float(v1)==0 else (float(v2)-float(v1))/float(v1)*100
    except: return 0.0

def ps(v): return f"{v:+.1f}%"

def color_cell(ws, row, col, val):
    if abs(val) > 15: ws.cell(row=row, column=col).fill = rfill
    elif abs(val) > 5: ws.cell(row=row, column=col).fill = yfill

# ===== SCENARIOS =====
scenarios = [
    {'title':'1组-风冷-乌兰D6B-27.5MW','label':'27.5MW 风冷 项目(乌兰D6-B)，工期30天',
     'input':{'total_mw':27.5,'total_duration':30,'cabinet_power':22,'total_cabinets':1500,'ac_type':'风冷','it_transformers':[['2.5',8]],'power_transformers':[['2.5',3]],'hybrid_transformers':[]},
     'man_staff':[('测试经理',1,28),('电气主测',1,28),('柴发主测',0,0),('暖通主测',2,25),('消防主测',0,0),('弱电主测',0,0),('电气测试员',3,23),('暖通测试员',0,0),('弱电测试员',1,23),('消防测试员',0,0),('记录员',27,20)],
     'man_tools':[('电能质量分析仪',8,'FLUKE 435'),('电能质量分析仪',1,'FLUKE 1775'),('热成像',9,'FLUKE Ti32'),('点温枪',1,'阈值750℃'),('钳形电流表',5,'FLUKE 381'),('温湿度仪',8,'FLUKE 971'),('万用表',6,'FLUKE 18B+'),('振动仪',1,'/'),('风速仪',2,'/'),('噪声仪',1,'/'),('电池内阻仪',2,'/')],
     'man_loads':[('风冷机架式假负载',0,0,'6KW/台'),('风冷机架式假负载',2823,18,'8KW/台'),('风冷集中式假负载',2,18,'500KW/台'),('风冷集中式假负载',0,0,'300KW/台'),('风冷高压柴发假负载',0,0,'2000KW/台')]},
    {'title':'2组-风液混合-太仓17-40MW','label':'40.0MW 风液混合 项目(太仓17#)，工期28天',
     'input':{'total_mw':40,'total_duration':28,'cabinet_power':[['液冷',30,400],['风冷',18,1000]],'total_cabinets':1400,'ac_type':'双冷源','it_transformers':[['2.5',12]],'power_transformers':[['2.5',4]],'hybrid_transformers':[['2.5',2]]},
     'man_staff':[('测试经理',1,28),('电气主测',1,28),('柴发主测',1,28),('暖通主测',1,28),('消防主测',1,28),('弱电主测',1,28),('电气测试员',25,28),('暖通测试员',10,28),('弱电测试员',4,28),('消防测试员',2,28),('记录员',1,28)],
     'man_tools':[('电能质量分析仪',10,'FLUKE 435'),('电能质量分析仪',1,'FLUKE 1775'),('热成像',10,'FLUKE Ti32'),('点温枪',1,'阈值750℃'),('钳形电流表',6,'FLUKE 381'),('温湿度仪',7,'FLUKE 971'),('万用表',4,'FLUKE 18B+'),('振动仪',1,'/'),('风速仪',3,'/'),('噪声仪',2,'/'),('电池内阻仪',2,'/')],
     'man_loads':[('风冷机架式假负载',1750,28,'8KW/台(含6kW折算)'),('液冷机架式假负载',400,28,'30KW/台'),('风冷集中式假负载',4,28,'500KW/台'),('风冷集中式假负载',2,28,'300KW/台'),('风冷高压柴发假负载',0,0,'2000KW/台')]},
    {'title':'3组-水冷-南通O楼-12MW','label':'12.0MW 水冷 项目(南通O楼)，工期32天',
     'input':{'total_mw':12,'total_duration':32,'cabinet_power':12,'total_cabinets':440,'ac_type':'水冷','it_transformers':[['1.6',6]],'power_transformers':[['1.3',2]],'hybrid_transformers':[],'project_type':'阿里巴拿马3.0'},
     # 南通O楼 阿里巴拿马3.0架构
     'man_staff':[('测试经理',1,32),('电气主测',2,32),('柴发主测',0,0),('暖通主测',2,32),('消防主测',1,32),('弱电主测',1,32),('电气测试员',5,32),('暖通测试员',2,32),('弱电测试员',2,32),('消防测试员',1,32),('记录员',0,0)],
     'man_tools':[('电能质量分析仪',5,'FLUKE 435'),('电能质量分析仪',0,'FLUKE 1775'),('热成像',8,'FLUKE Ti32'),('点温枪',10,'阈值750℃'),('钳形电流表',4,'FLUKE 381'),('温湿度仪',10,'FLUKE 971'),('万用表',4,'FLUKE 18B+'),('振动仪',2,'/'),('风速仪',4,'/'),('噪声仪',2,'/'),('电池内阻仪',2,'/')],
     'man_loads':[('风冷机架式假负载',420,32,'6KW/台'),('风冷机架式假负载',1440,32,'8KW/台'),('液冷机架式假负载',0,0,'30KW/台'),('风冷集中式假负载',2,30,'500KW/台'),('风冷集中式假负载',0,0,'300KW/台'),('风冷高压柴发假负载',0,0,'2000KW/台')]},
    {'title':'4组-液冷-30MW-28天','label':'30.0MW 液冷 项目(手算)，工期28天',
     'input':{'total_mw':30,'total_duration':28,'cabinet_power':30,'total_cabinets':1000,'ac_type':'水冷','it_transformers':[['2.5',10]],'power_transformers':[['2.5',4]],'hybrid_transformers':[['2.5',1]]},
     'man_staff':[('测试经理',1,28),('电气主测',1,28),('柴发主测',1,28),('暖通主测',1,28),('消防主测',1,28),('弱电主测',1,28),('电气测试员',21,24),('暖通测试员',12,28),('弱电测试员',7,28),('消防测试员',1,28),('记录员',5,28)],
     'man_tools':[('电能质量分析仪',14,'FLUKE 435'),('电能质量分析仪',2,'FLUKE 1775'),('热成像',7,'FLUKE Ti32'),('点温枪',4,'阈值750℃'),('钳形电流表',3,'FLUKE 381'),('温湿度仪',8,'FLUKE 971'),('万用表',7,'FLUKE 18B+'),('振动仪',2,'/'),('风速仪',4,'/'),('噪声仪',2,'/'),('电池内阻仪',2,'/')],
     'man_loads':[('液冷机架式假负载',306,28,'30KW/台'),('风冷机架式假负载',0,0,'6KW/台'),('风冷机架式假负载',0,0,'8KW/台'),('风冷集中式假负载',5,26,'500KW/台'),('风冷集中式假负载',0,0,'300KW/台'),('风冷高压柴发假负载',0,0,'2000KW/台')]},
]

# ===== BUILD WORKBOOK =====
wb = openpyxl.Workbook()
wb.remove(wb.active)

def write_staff_full(ws, row, manual_staff, model_staff, dur):
    """Personnel with quantity + days + man-days"""
    for i, (name, m_qty, m_days) in enumerate(manual_staff):
        m_md = m_qty * m_days
        c_qty, c_days = 0, dur
        for cname, cq, cd in model_staff:
            if cname == name:
                c_qty, c_days = cq, cd
                break
        c_md = c_qty * c_days

        dq = pct(m_qty, c_qty) if m_qty else 0
        dd = pct(m_days, c_days) if m_days else 0
        dm = pct(m_md, c_md) if m_md else 0

        row_data = [i+1, name, m_qty, c_qty, ps(dq), m_days, c_days, ps(dd), m_md, c_md, ps(dm)]
        for j, v in enumerate(row_data, 1):
            ws.cell(row=row, column=j, value=v).border = border
        for col, v in [(5,dq),(8,dd),(11,dm)]:
            color_cell(ws, row, col, v)
        row += 1

    # Totals
    m_tq = sum(r[1] for r in manual_staff)
    c_tq = sum(r[1] for r in model_staff)
    m_td = sum(r[2] for r in manual_staff)
    c_td = sum(r[2] for r in model_staff)
    m_md = sum(r[1]*r[2] for r in manual_staff)
    c_md = sum(r[1]*r[2] for r in model_staff)
    for col, mv, cv in [(3,m_tq,c_tq),(6,m_td,c_td),(9,m_md,c_md)]:
        ws.cell(row=row, column=col, value=mv).border = border
        ws.cell(row=row, column=col+1, value=cv).border = border
        d = pct(mv, cv) if mv else 0
        ws.cell(row=row, column=col+2, value=ps(d)).border = border
    ws.cell(row=row, column=1, value='').border = border
    ws.cell(row=row, column=2, value='合计').border = border
    ws.cell(row=row, column=2).font = Font(bold=True)
    return row + 2

def write_tools_full(ws, row, title, manual_data, model_data, dur, is_load=False):
    """Tools/loads with quantity + days + 台天"""
    ws.cell(row=row, column=1, value=title).font = hfont
    ws.merge_cells(f'A{row}:L{row}')
    row += 1
    hdrs = ['序号','名称',
            '手算\n台数','计算器\n台数','台数\n偏差率',
            '手算\n天数','计算器\n天数','天数\n偏差率',
            '手算\n台天','计算器\n台天','台天\n偏差率','规格']
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=i, value=h); c.font = hfont; c.border = border
        c.alignment = Alignment(wrap_text=True)
    row += 1

    tm_q = tc_q = tm_d = tc_d = tm_t = tc_t = 0
    for i, m_item in enumerate(manual_data):
        name = m_item[0]; m_q = int(m_item[1]) if m_item[1] else 0
        if is_load:
            m_d = int(m_item[2]) if len(m_item) > 2 and m_item[2] else 0
            spec = str(m_item[3]) if len(m_item) > 3 else ''
        else:
            m_d = dur
            spec = str(m_item[2]) if len(m_item) > 2 else ''
        m_t = m_q * m_d

        c_q = c_d = 0
        if i < len(model_data):
            c_q = int(model_data[i][1]) if model_data[i][1] else 0
            c_d = int(model_data[i][2]) if len(model_data[i]) > 2 and model_data[i][2] else (dur if not is_load else 0)
        c_t = c_q * c_d

        dq = pct(m_q, c_q) if m_q else 0
        dd = pct(m_d, c_d) if m_d else 0
        dt = pct(m_t, c_t) if m_t else 0

        rv = [i+1, name, m_q, c_q, ps(dq), m_d, c_d, ps(dd), m_t, c_t, ps(dt), spec]
        for j, v in enumerate(rv, 1):
            ws.cell(row=row, column=j, value=v).border = border
        for col, v in [(5,dq),(8,dd),(11,dt)]:
            color_cell(ws, row, col, v)
        tm_q += m_q; tc_q += c_q; tm_d += m_d; tc_d += c_d; tm_t += m_t; tc_t += c_t
        row += 1

    for col, mv, cv in [(3,tm_q,tc_q),(6,tm_d,tc_d),(9,tm_t,tc_t)]:
        ws.cell(row=row, column=col, value=mv).border = border
        ws.cell(row=row, column=col+1, value=cv).border = border
        d = pct(mv, cv) if mv else 0
        ws.cell(row=row, column=col+2, value=ps(d)).border = border
    ws.cell(row=row, column=1, value='').border = border
    ws.cell(row=row, column=2, value='合计').border = border
    ws.cell(row=row, column=2).font = Font(bold=True)
    return row + 2

# Summary sheet
ws_sum = wb.create_sheet('总结')
ws_sum.cell(row=1,column=1,value='资源计算器校准偏差报告 —— 2026.6.18').font = tfont
ws_sum.merge_cells('A1:L1')
sr = 3
for i, sc in enumerate(scenarios):
    r = rp.calculate(rp.make_input(sc['input']), cfg)
    mq = sum(s[1] for s in sc['man_staff'])
    cq = r['汇总']['峰值同时在场']
    mmd = sum(s[1]*s[2] for s in sc['man_staff'])
    cmd = r['汇总']['总人天']
    ws_sum.cell(row=sr, column=1, value=f"{i+1}、{sc['label']} | 类型:{r['项目信息']['项目类型']} | 峰值偏差 {ps(pct(mq,cq))} | 人天偏差 {ps(pct(mmd,cmd))}").font = Font(bold=True)
    sr += 2

# Project sheets
for sc in scenarios:
    ws = wb.create_sheet(sc['title'])
    inp = rp.make_input(sc['input']); r = rp.calculate(inp, cfg)
    dur = inp['dur']; pt = r['项目信息']['项目类型']
    it=r['IT链路']; pw=r['动力链路']; hb=r['混合链路']; hv=r['暖通']
    gen=r['柴发']; wk=r['弱电']; fi=r['消防']; fx=r['固定人员']
    et=r['工器具']['电气工器具']; ht=r['工器具']['暖通工器具']; ld=r['负载']

    ws.cell(row=1,column=1,value=f'资源计算器校准偏差报告 —— {sc["title"]}').font = tfont
    ws.merge_cells('A1:L1')
    ws.cell(row=2,column=1,value=f'参考项目：{sc["label"]}  |  基准：手算/实际  |  对比：资源计算器').font = Font(size=10)
    ws.merge_cells('A2:L2')

    # Model staff with days
    elec_on = it['在场']+pw['在场']+hb['在场']
    elec_dur = max(it.get('实际工期',0), pw.get('实际工期',0), hb.get('实际工期',0)) or dur
    hvac_pk = max(hv['功能测试']['在场'], hv['场景压测']['在场'], hv.get('前端冷源',{}).get('人数',0), hv.get('安装检查',{}).get('人数',0))
    proj_type_cfg = cfg.get('project_type_config', {}).get(pt, {})
    skip_rec = proj_type_cfg.get('skip_recorders', False)
    wk_testers = wk.get('电气记录员', wk['小计']-wk.get('主测',1))
    wk_recorders = 0 if skip_rec else wk.get('暖通记录员',0)
    mod_staff = [('测试经理',fx.get('项目经理',1),dur),('电气主测',fx.get('电气主测',1),dur),('柴发主测',gen['主测'],dur),('暖通主测',fx.get('暖通主测',1),dur),('消防主测',fi['主测'],dur),('弱电主测',wk['主测'],dur),('电气测试员',elec_on,elec_dur),('暖通测试员',hvac_pk,dur),('弱电测试员',wk_testers,dur),('消防测试员',fi.get('测试员',fi['小计']-fi.get('主测',1)),dur),('记录员',gen.get('记录员',0)+wk_recorders,dur)]

    # Model tools
    m_tools={'电能质量分析仪':et.get('电能质量分析仪435',0),'电能质量分析仪2':et.get('电能质量分析仪1775',1),'热成像':et.get('热成像',0)+ht.get('热成像_暖通',0),'点温枪':et.get('点温枪',1),'钳形电流表':et.get('钳形电流表381',0),'钳形电流表2':et.get('钳形电流表319',0),'温湿度仪':et.get('温湿度仪971',0)+ht.get('温湿度仪971_暖通',0),'万用表':et.get('万用表',0),'振动仪':et.get('振动仪',1),'风速仪':ht.get('风速仪',0),'噪声仪':et.get('噪声仪',1),'电池内阻仪':et.get('电池内阻仪',1)}
    mod_tools = []; used = set()
    for mt in sc['man_tools']:
        name = mt[0]; qty = 0
        for k, v in m_tools.items():
            if k.startswith(name) and k not in used and v >= 0:
                qty = v; used.add(k); break
        mod_tools.append((name, qty, dur))

    # Model loads
    p30=ld.get('30kW',{}).get('总需求',0); p6=ld.get('6kW',{}).get('总需求',0); p8=ld.get('8kW',{}).get('总需求',0)
    p500=ld.get('500kW',{}).get('总需求',0); p300=ld.get('300kW',{}).get('总需求',0)
    air_dur = dur - 6  # 风冷假负载测试天数（扣安装检查）
    liq_dur = max(math.ceil(dur / 6), 4)  # 液冷负载测试窗口较短
    if pt == '风冷':
        mod_loads = [('风冷机架式假负载',p6,air_dur if p6 else 0),('风冷机架式假负载',p8,air_dur if p8 else 0),('风冷集中式假负载',p500,air_dur if p500 else 0),('风冷集中式假负载',p300,air_dur if p300 else 0),('风冷高压柴发假负载',0,0)]
    elif pt == '风液混合':
        mod_loads = [('风冷机架式假负载',p8,air_dur if p8 else 0),('液冷机架式假负载',p30,liq_dur if p30 else 0),('风冷集中式假负载',p500,air_dur if p500 else 0),('风冷集中式假负载',p300,air_dur if p300 else 0),('风冷高压柴发假负载',0,0)]
    else:
        mod_loads = [('液冷机架式假负载',p30,air_dur if p30 else 0),('风冷机架式假负载',p6,air_dur if p6 else 0),('风冷机架式假负载',p8,air_dur if p8 else 0),('风冷集中式假负载',p500,air_dur if p500 else 0),('风冷集中式假负载',p300,air_dur if p300 else 0),('风冷高压柴发假负载',0,0)]

    # Write
    row = 4
    row = write_staff_full(ws, row+2, sc['man_staff'], mod_staff, dur)
    # 工器具：展示用前3项，合计行汇总用全项
    row = write_tools_full(ws, row, '2、工器具清单（仅展示前3项，合计行按全16项汇总）', sc['man_tools'][:3], mod_tools[:3], dur)
    # 手动更新合计行为全项的汇总
    full_mq = sum(m[1] for m in sc['man_tools'])
    full_cq = sum(m[1] for m in mod_tools)
    ws.cell(row=row-1, column=3, value=full_mq)
    ws.cell(row=row-1, column=4, value=full_cq)
    full_dev = (full_cq-full_mq)/full_mq*100 if full_mq else 0
    ws.cell(row=row-1, column=5, value=f'{full_dev:+.1f}%')
    # also update days and 台天 totals
    full_md = sum(m[1]*dur for m in sc['man_tools'])
    full_cd = sum(m[1]*dur for m in mod_tools)
    ws.cell(row=row-1, column=6, value=full_md)
    ws.cell(row=row-1, column=7, value=dur)
    dd = (dur-dur)/dur*100 if dur else 0
    ws.cell(row=row-1, column=8, value=f'{dd:+.1f}%')
    ws.cell(row=row-1, column=9, value=full_md*dur)
    ws.cell(row=row-1, column=10, value=full_cd*dur)
    ws.cell(row=row-1, column=11, value=f'{full_dev:+.1f}%')
    row = write_tools_full(ws, row, '3、假负载清单', sc['man_loads'], mod_loads, dur, is_load=True)

    # 4-6
    row += 1; ws.cell(row=row,column=1,value='4、劳务清单').font = hfont; ws.merge_cells(f'A{row}:L{row}')
    row += 2; ws.cell(row=row,column=2,value='投入人天').border = border; ws.cell(row=row,column=3,value=0).border = border; ws.cell(row=row,column=4,value=0).border = border; ws.cell(row=row,column=5,value='0%').border = border
    row += 3; ws.cell(row=row,column=1,value='5、机柜及PDU清单').font = hfont; ws.merge_cells(f'A{row}:L{row}')
    for nm in ['机柜','PDU']:
        row += 2; ws.cell(row=row,column=2,value=nm).border = border; ws.cell(row=row,column=3,value=0).border = border; ws.cell(row=row,column=4,value=0).border = border; ws.cell(row=row,column=5,value='0%').border = border
    row += 3; ws.cell(row=row,column=1,value='6、参考项目依据').font = hfont; ws.merge_cells(f'A{row}:L{row}')
    row += 2; ws.cell(row=row,column=2,value=f'参考项目：{sc["label"]}').border = border; ws.merge_cells(f'B{row}:L{row}')

    for c,w in {1:6,2:18,3:10,4:10,5:12,6:10,7:10,8:12,9:10,10:10,11:12,12:25}.items():
        ws.column_dimensions[get_column_letter(c)].width = w

# Save
out_dir = r'D:\世纪互联\工作文档\read\2026.6.18资源计算器校准'
os.makedirs(out_dir, exist_ok=True)
out_path = os.path.join(out_dir, '偏差报告_校准对比_完整版.xlsx')
try:
    wb.save(out_path)
except PermissionError:
    out_path = os.path.join(out_dir, '偏差报告_校准对比_v6.xlsx')
    wb.save(out_path)
print(f"Saved: {out_path}")
