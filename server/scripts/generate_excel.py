#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据中心测试资源规划 Excel 报告生成器 V200
==============================================
基于模板.xlsx 格式输出，保留模板原有名称，只填充数值。

版本历史：
  V200 (2026-05-27)
    - 版本号升级，与 resource_plan.py V200 配套

  V100 (2026-05-25)
    - 修复浮点数 ceil 精度问题
    - 劳务清单和机柜PDU清单默认填0
    
  依赖：
    - openpyxl
    - templates/*.xlsx
"""

import json
import math
import sys
from openpyxl import load_workbook

__version__ = "200.0.0"


def create_template_excel(result: dict, template_path: str, output_path: str):
    """根据计算结果填充模板Excel

    参数:
        result: calculate() 返回的结果字典
        template_path: 模板文件路径
        output_path: 输出文件路径
    """
    wb = load_workbook(template_path)
    ws = wb.active

    # ---- 从结果中提取数据 ----
    duration = int(result["项目信息"]["工期"].replace("天", ""))
    it = result["IT链路"]
    pw = result["动力链路"]
    hb = result.get("混合链路", {"在场": 0, "实际工期": 0})
    hvac = result["暖通"]
    weak = result["弱电"]
    fire = result["消防"]
    gen = result["柴发"]
    loads = result["负载"]

    # ---- 电气综合数据 ----
    elec_on_site = it["在场"] + pw["在场"] + hb["在场"]
    elec_dur = max(
        it.get("实际工期", 0),
        pw.get("实际工期", 0),
        hb.get("实际工期", 0),
    )
    if elec_dur == 0:
        elec_dur = duration

    hvac_peak = max(
        hvac["功能测试"]["在场"],
        hvac["场景压测"]["在场"],
        hvac.get("前端冷源", {}).get("人数", 0),
        hvac.get("安装检查", {}).get("人数", 0),
    )

    # ========== 1. 人员投入清单 ==========
    staff = [
        (3, 1, duration),                                # 测试经理
        (4, 1, duration),                                # 电气主测
        (5, gen["主测"], duration),                      # 柴发主测
        (6, 1, duration),                                # 暖通主测
        (7, fire["主测"], duration),                     # 消防主测
        (8, weak["主测"], duration),                     # 弱电主测
        (9, elec_on_site, elec_dur),                     # 电气测试员
        (10, hvac_peak, duration),                       # 暖通测试员
        (11, weak["电气记录员"], duration),              # 弱电测试员
        (12, fire["测试员"], duration),                  # 消防测试员
        (13, gen["记录员"] + weak["暖通记录员"], duration),  # 记录员
    ]
    for row, count, days in staff:
        ws.cell(row=row, column=3, value=count)
        ws.cell(row=row, column=4, value=days)
        ws.cell(row=row, column=5, value=f"=C{row}*D{row}")

    # ========== 2. 工器具清单 ==========
    tool_cnt = {
        18: 6, 19: 2, 20: 2, 21: 4, 22: 6, 23: 6,
        24: 10, 25: 10, 26: 2, 27: 4, 28: 6,
        29: 2, 30: 2, 31: 2, 32: 2, 33: 3,
    }
    for row, cnt in tool_cnt.items():
        ws.cell(row=row, column=3, value=cnt)
        ws.cell(row=row, column=4, value=duration)
        ws.cell(row=row, column=5, value=f"=C{row}*D{row}")

    # ========== 3. 假负载清单 ==========
    spare = 0.1
    l6 = loads["6kW"]["总需求"]
    l8 = loads["8kW"]["总需求"]
    l5 = loads["500kW"]["总需求"]
    l3 = loads["300kW"]["总需求"]

    # 行38: 6KW机架式（ceil修正浮点精度）
    ws.cell(row=38, column=3, value=math.ceil(l6 * (1 + spare) - 1e-12))
    ws.cell(row=38, column=4, value=duration)
    ws.cell(row=38, column=5, value="=C38*D38")

    # 行39: 8KW机架式
    ws.cell(row=39, column=3, value=math.ceil(l8 * (1 + spare) - 1e-12))
    ws.cell(row=39, column=4, value=duration)
    ws.cell(row=39, column=5, value="=C39*D39")

    # 行40: 500KW集中式（无冗余）
    ws.cell(row=40, column=3, value=l5)
    ws.cell(row=40, column=4, value=duration)
    ws.cell(row=40, column=5, value="=C40*D40")

    # 行41: 300KW集中式（无冗余）
    ws.cell(row=41, column=3, value=l3)
    ws.cell(row=41, column=4, value=duration)
    ws.cell(row=41, column=5, value="=C41*D41")

    # 行42: 2000KW柴发（默认填0）
    ws.cell(row=42, column=3, value=0)
    ws.cell(row=42, column=4, value=duration)
    ws.cell(row=42, column=5, value="=C42*D42")

    # ========== 4. 劳务清单（默认填0）==========
    ws.cell(row=47, column=2, value=0)
    ws.cell(row=47, column=3, value=0)

    # ========== 5. 机柜及PDU清单（默认填0）==========
    ws.cell(row=51, column=3, value=0)
    ws.cell(row=51, column=4, value=0)
    ws.cell(row=51, column=5, value=0)
    ws.cell(row=52, column=3, value=0)
    ws.cell(row=52, column=4, value=0)
    ws.cell(row=52, column=5, value=0)

    # ========== 6. 参考项目依据 ==========
    ws.cell(
        row=56, column=2,
        value=(
            f"参考项目：{result['项目信息']['总容量']} "
            f"{result['项目信息']['空调']} 项目，"
            f"工期{result['项目信息']['工期']}"
        ),
    )

    wb.save(output_path)
    print(f"✅ 模板格式Excel已生成: {output_path}")
    print(f"   总人天: {result['汇总']['总人天']}, "
          f"峰值: {result['汇总']['峰值同时在场']}")


def main():
    if len(sys.argv) < 4:
        print("用法: python generate_excel.py <result.json> <模板.xlsx> <output.xlsx>")
        sys.exit(1)

    with open(sys.argv[1], "r", encoding="utf-8") as f:
        data = json.load(f)

    create_template_excel(data, sys.argv[2], sys.argv[3])


if __name__ == "__main__":
    main()
