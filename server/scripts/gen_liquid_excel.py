#!/usr/bin/env python3
"""液冷项目 Excel 输出 — 匹配参考模板格式"""

import json, sys, os, math
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import resource_plan as rp

with open(os.path.join(os.path.dirname(__file__), '..', 'config_v100.json'), encoding='utf-8') as f:
    cfg = json.load(f)

# ====== 液冷项目计算 ======
d = {'total_mw': 30, 'total_duration': 28,
     'cabinet_power': 30, 'total_cabinets': 1000,
     'ac_type': '水冷',
     'it_transformers': [['2.5', 10]], 'power_transformers': [['2.5', 4]],
     'hybrid_transformers': [['2.5', 1]]}
inp = rp.make_input(d)
r = rp.calculate(inp, cfg)

dur = 28
it, pw, hb = r['IT链路'], r['动力链路'], r['混合链路']
hv, gen, wk, fi, fx = r['暖通'], r['柴发'], r['弱电'], r['消防'], r['固定人员']
loads = r['负载']
tools = r['工器具']
peak = r['汇总']['峰值同时在场']
md = r['汇总']['总人天']

elec_on = it['在场'] + pw['在场'] + hb['在场']
elec_dur = max(it.get('实际工期', 0), pw.get('实际工期', 0), hb.get('实际工期', 0)) or dur
hvac_peak = max(hv['功能测试']['在场'], hv['场景压测']['在场'],
                hv.get('前端冷源', {}).get('人数', 0), hv.get('安装检查', {}).get('人数', 0))

et = tools['电气工器具']
ht = tools['暖通工器具']
elec_groups = it['并行数'] + pw['并行数'] + hb['并行数']

print(f"液冷 30MW 28天 | Type={r['项目信息']['项目类型']}")
print(f"Peak={peak} MD={md}")
print(f"IT={it['在场']}({it['并行数']}p×{it['每台人数']}pp) PW={pw['在场']}({pw['并行数']}p) HB={hb['在场']}({hb['并行数']}p)")
print(f"HVACpk={hvac_peak} Gen={gen['小计']} Wk={wk['小计']} Fi={fi['小计']} Fx={fx['小计']}")
print(f"elec_on={elec_on} elec_groups={elec_groups}")
print(f"30kW={loads['30kW']['总需求']} 500kW={loads['500kW']['总需求']}")

# ====== 生成 Excel（基于参考文件模板）======
import openpyxl
from copy import copy

ref_path = r'D:\世纪互联\工作文档\read\2026.5.27资源计算器校准\手算资源规划_组3_30MW_28天.xlsx'
wb = openpyxl.load_workbook(ref_path)
ws = wb.active
ws.title = '1、液冷项目'

# ====== 1. 人员投入清单 (rows 3-13, 参考模板) ======
# 参考模板 row 2=header, row 3-13=data, row 14=合计
staff = [
    (3,  1, dur),                                            # 测试经理
    (4,  1, dur),                                            # 电气主测
    (5,  gen['主测'], dur),                                   # 柴发主测
    (6,  1, dur),                                            # 暖通主测
    (7,  fi['主测'], dur),                                    # 消防主测
    (8,  wk['主测'], dur),                                    # 弱电主测
    (9,  elec_on, elec_dur),                                 # 电气测试员
    (10, hvac_peak, dur),                                    # 暖通测试员
    (11, wk.get('电气记录员', wk['小计'] - wk.get('主测', 1)), dur),   # 弱电测试员
    (12, fi.get('测试员', fi['小计'] - fi.get('主测', 1)), dur),      # 消防测试员
    (13, gen.get('记录员', 0) + wk.get('暖通记录员', 0), dur),         # 记录员
]
for row, count, days in staff:
    ws.cell(row=row, column=3, value=count)
    ws.cell(row=row, column=4, value=days)
    ws.cell(row=row, column=5, value=f'=C{row}*D{row}')

# 合计行 (row 14)
ws.cell(row=14, column=3, value=peak)
ws.cell(row=14, column=4, value='-')
ws.cell(row=14, column=5, value=md)

# ====== 2. 工器具清单 (rows 18-33, 参考模板) ======
# Standard tool counts per project scale (scaled by elec_groups, max reasonable values)
n = elec_groups  # number of electrical groups
tool_data = [
    (18, max(n * 2, 4), 'FLUKE 435', '配套6000A电流环；要求435-2'),
    (19, 2, 'FLUKE 1775', '至少2000A以上电流环'),
    (20, max(n, 2), 'FLUKE Ti32', ''),
    (21, 4, '阈值750℃', ''),
    (22, max(n, 4), '/', ''),
    (23, max(n, 4), '/', ''),
    (24, 10, '16A', 'PDU欧标'),
    (25, 10, '10A', 'PDU欧标'),
    (26, max(n // 2, 2), 'FLUKE 381', '大线圈，量程1500~2000A'),
    (27, max(ht.get('温湿度仪971', 0) // 2, 4), 'FLUKE 971', ''),
    (28, max(n, 4), 'FLUKE 18B+', ''),
    (29, 2, '/', ''),
    (30, max(ht.get('风速仪', 0) // 2, 2), '/', ''),
    (31, 2, '/', ''),
    (32, 2, '福禄克、日置', ''),
    (33, max(n * 2, 3), '/', ''),
]
for row, cnt, model, desc in tool_data:
    ws.cell(row=row, column=3, value=cnt)
    ws.cell(row=row, column=4, value=dur)
    ws.cell(row=row, column=5, value=f'=C{row}*D{row}')
    if model and ws.cell(row=row, column=6).value in (None, '/', '-'):
        ws.cell(row=row, column=6, value=model)
    if desc:
        ws.cell(row=row, column=7, value=desc)

# ====== 3. 假负载清单 (rows 38-42) ======
s = 0.1
p30 = loads.get('30kW', {}).get('总需求', 0)
p6 = loads.get('6kW', {}).get('总需求', 0)
p8 = loads.get('8kW', {}).get('总需求', 0)
p500 = loads.get('500kW', {}).get('总需求', 0)
p300 = loads.get('300kW', {}).get('总需求', 0)

# Row 38: 液冷机架式假负载 (replaces 6kW row)
ws.cell(row=38, column=1, value=1)
ws.cell(row=38, column=2, value='液冷机架式假负载')
ws.cell(row=38, column=3, value=math.ceil(p30 * (1 + s) - 1e-12))
ws.cell(row=38, column=4, value=dur)
ws.cell(row=38, column=5, value='=C38*D38')
ws.cell(row=38, column=6, value=s)
ws.cell(row=38, column=7, value='30KW/台（含FD83快插及软管和卡盘）')

# Row 39: 风冷机架式6kW (0 for pure liquid)
ws.cell(row=39, column=1, value=2)
ws.cell(row=39, column=2, value='风冷机架式假负载')
ws.cell(row=39, column=3, value=math.ceil(p6 * (1 + s) - 1e-12) if p6 > 0 else 0)
ws.cell(row=39, column=4, value=dur)
ws.cell(row=39, column=5, value='=C39*D39')
ws.cell(row=39, column=6, value=s)
ws.cell(row=39, column=7, value='6KW/台')

# Row 40: 风冷机架式8kW (0 for pure liquid)
ws.cell(row=40, column=1, value=3)
ws.cell(row=40, column=2, value='风冷机架式假负载')
ws.cell(row=40, column=3, value=math.ceil(p8 * (1 + s) - 1e-12) if p8 > 0 else 0)
ws.cell(row=40, column=4, value=dur)
ws.cell(row=40, column=5, value='=C40*D40')
ws.cell(row=40, column=6, value=s)
ws.cell(row=40, column=7, value='8KW/台')

# Row 41: 集中式500kW
ws.cell(row=41, column=3, value=p500)
ws.cell(row=41, column=4, value=dur - 2 if p500 > 0 else 0)
ws.cell(row=41, column=5, value='=C41*D41')

# Row 42: 集中式300kW
ws.cell(row=42, column=3, value=p300)
ws.cell(row=42, column=4, value=dur - 2 if p300 > 0 else 0)
ws.cell(row=42, column=5, value='=C42*D42')

# Row 43: 柴发2000kW
ws.cell(row=43, column=3, value=0)
ws.cell(row=43, column=4, value=dur)
ws.cell(row=43, column=5, value='=C43*D43')

# ====== 4. 劳务清单 (row 47) ======
ws.cell(row=47, column=2, value=0)
ws.cell(row=47, column=3, value=0)

# ====== 5. 机柜PDU (rows 51-52) ======
ws.cell(row=51, column=3, value=0)
ws.cell(row=52, column=3, value=0)

# ====== 6. 参考项目 (row 56) ======
ws.cell(row=56, column=2,
        value=f'参考项目：{r["项目信息"]["总容量"]} 液冷 项目（水冷空调+液冷30kW机柜），工期{r["项目信息"]["工期"]}')

# Save
out_path = os.path.join(os.path.dirname(__file__), '..', 'test6', 'output', '液冷_30MW_28天_资源规划.xlsx')
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f'\nDone: {out_path}')
